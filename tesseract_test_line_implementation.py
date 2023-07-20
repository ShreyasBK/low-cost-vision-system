#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import cv2
import matplotlib.pyplot as plt
import numpy as np
get_ipython().run_line_magic('matplotlib', 'inline')
import pytesseract
import pandas as pd
import openpyxl
from datetime import datetime
import os

# Load an image using cv2.imread()
img = cv2.imread(r'C:\Users\SHREYAS.BK\MVviewer\pictures\Pic_2023_01_23_175736_2.bmp')

img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)



# Define the crop region
region_width = 950
region_height = 300
region_center_x = 1300
region_center_y = 1000


# Crop the image
cropped_img = cv2.getRectSubPix(img_rgb, (region_width, region_height), (region_center_x, region_center_y))


# Define the kernel for the morphological transformation
kernel = np.ones((5,5), np.uint8)


# Apply erosion to the image
erosion = cv2.erode(cropped_img, kernel, iterations=1)

# Apply a binary threshold to the image
ret,thresh = cv2.threshold(erosion, 180, 255, cv2.THRESH_BINARY)
# plt.imshow(img_rgb,cmap='gray')
# plt.show()
plt.imshow(thresh,cmap='gray')
plt.show()

# Load the existing Excel file
wb = openpyxl.load_workbook(r'D:\Python\OCR_Recognition_codes\Various_detection_methods\sample.xlsx')

# Get the active sheet in the workbook
ws = wb.active

# Run OCR on the image using pytesseract.image_to_string()
text = pytesseract.image_to_string(thresh, config='--psm 6')
print(text)
# Store the OCR text in a variable
a = text

# Find the next empty row in column A
next_row = ws.max_row + 1

# Write the value of variable 'a' to the next empty cell in column A
ws.cell(row=next_row, column=1).value = a

# Add the headers for the new columns
ws.cell(row=1, column=2).value = "Date"
ws.cell(row=1, column=3).value = "Time"

# Get the current date and time
now = datetime.now()

# Write the current date to the next empty cell in column B
ws.cell(row=next_row, column=2).value = now.date()

# Write the current time to the next empty cell in column C
ws.cell(row=next_row, column=3).value = now.time()

# Save the workbook
wb.save(r'D:\Python\OCR_Recognition_codes\Various_detection_methods\sample.xlsx')
timestamp = now.strftime("%Y-%m-%d_%H-%M-%S")
filename = "image_{}.bmp".format(timestamp)
cv2.imwrite(r"D:\Python\MV_viewer_images\{}.bmp".format(filename), thresh)



# In[ ]:





# In[ ]:




